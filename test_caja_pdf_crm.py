"""
Pruebas para las tres funcionalidades nuevas pedidas por Wilmer:

  A. Cuadre de caja real (base inicial + conteo físico + diferencia),
     con recálculo del lado del servidor (no se puede manipular desde
     el navegador) y registro en auditoría, incluyendo re-cierres.
  B. PDF de tiquetes del día: una copia fiel (mismo contenido que
     templates/ticket.html) de cada tiquete de entrada y de salida
     generado en una fecha, agrupados por vehículo.
  C. Excel de clientes (CRM): placa, teléfono, tipo (Mensualidad/
     Diario) y fecha de vencimiento (solo si es mensualidad).
"""
import io
import os
import re

os.environ["DEVELOPMENT"] = "true"
os.environ["PRINTER_API_KEY"] = "test_key_caja_pdf_crm"
import app as appmod
from openpyxl import load_workbook

try:
    import pdfplumber
    TIENE_PDFPLUMBER = True
except ImportError:
    TIENE_PDFPLUMBER = False

client = appmod.app.test_client()


def limpiar_db():
    conn = appmod.get_db()
    conn.execute("DELETE FROM parqueadero")
    conn.execute("DELETE FROM mensualidades")
    conn.execute("DELETE FROM clientes_frecuentes")
    conn.execute("DELETE FROM cierres_caja")
    conn.execute("DELETE FROM auditoria")
    conn.execute("UPDATE consecutivo SET numero = 0 WHERE id = 1")
    conn.commit()
    conn.close()


def set_sesion(usuario, rol):
    with client.session_transaction() as sess:
        sess["usuario"] = usuario
        sess["rol"] = rol


def csrf_from(html):
    m = re.search(r'name="csrf_token" value="([^"]+)"', html)
    if not m:
        m = re.search(r'name="csrf-token" content="([^"]+)"', html)
    assert m, "No se encontró csrf_token en la página -- ¿cambió el HTML?"
    return m.group(1)


def get_con_token(url):
    r = client.get(url)
    return r, csrf_from(r.get_data(as_text=True))


def texto_pdf(contenido_bytes):
    assert TIENE_PDFPLUMBER, "Instala pdfplumber para poder leer el contenido del PDF en la prueba"
    texto_total = ""
    with pdfplumber.open(io.BytesIO(contenido_bytes)) as pdf:
        for pagina in pdf.pages:
            texto_total += (pagina.extract_text() or "") + "\n"
    return texto_total


def crear_entrada_salida(placa, tipo, valor, metodo_pago="Efectivo", fecha_entrada=None, fecha_salida=None, marca="MarcaX"):
    """
    Inserta directamente en la BD un registro de entrada+salida ya
    completo (evita pasar por las rutas web, más rápido para armar
    escenarios de prueba con fechas específicas).
    """
    conn = appmod.get_db()
    ticket_num = conn.execute("SELECT numero FROM consecutivo WHERE id=1").fetchone()["numero"] + 1
    hoy_str = appmod.get_colombia_time().strftime("%Y-%m-%d %H:%M:%S")
    he = fecha_entrada or hoy_str
    hs = fecha_salida or (hoy_str if fecha_salida is not False else None)
    conn.execute("""
        INSERT INTO parqueadero (placa, tipo, hora_entrada, hora_salida, ticket_num, marca, celular, cajero, valor, metodo_pago)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (placa, tipo, he, hs, ticket_num, marca, "3000000000", "admin", valor, metodo_pago))
    conn.execute("UPDATE consecutivo SET numero=? WHERE id=1", (ticket_num,))
    conn.commit()
    conn.close()
    return ticket_num


limpiar_db()
set_sesion("admin", "admin")

# ═════════════════════════════════════════════════════════════════
# PRUEBA A: CUADRE DE CAJA (conteo físico real)
# ═════════════════════════════════════════════════════════════════

hoy = appmod.get_colombia_time().strftime("%Y-%m-%d")

# A1. Sin ningún cierre previo, el formulario sugiere base=0
r = client.get("/caja/cerrar")
assert r.status_code == 200
assert "Cerrar caja de este día" in r.get_data(as_text=True)
print("PRUEBA A1: OK -- sin cierres previos, se muestra el formulario para cerrar por primera vez")

# A2. Registramos una venta en efectivo de $7.000 hoy
crear_entrada_salida("CAJA001", "moto", 7000, metodo_pago="Efectivo")

conn = appmod.get_db()
efectivo_ventas = appmod._efectivo_ventas_dia(conn, hoy)
conn.close()
assert efectivo_ventas == 7000
print("PRUEBA A2: OK -- _efectivo_ventas_dia() suma correctamente las ventas en efectivo del día")

# A3. Cerramos caja: base 50.000 + venta 7.000 = 57.000 esperado.
#     Contamos exactamente eso -> diferencia debe ser 0 (cuadrada).
_, tk = get_con_token("/caja/cerrar")
r = client.post("/caja/cerrar", data={
    "fecha": hoy, "base_inicial": "50000", "efectivo_contado": "57000", "observaciones": "prueba cuadrada",
    "csrf_token": tk,
}, follow_redirects=True)
assert r.status_code == 200
html = r.get_data(as_text=True)
assert "cuadrada exactamente" in html.lower() or "Sin diferencia" in html
print("PRUEBA A3: OK -- cuadre exacto (base + ventas = contado) se guarda con diferencia 0")

conn = appmod.get_db()
cierre = conn.execute("SELECT * FROM cierres_caja WHERE fecha=?", (hoy,)).fetchone()
conn.close()
assert cierre is not None
assert cierre["base_inicial"] == 50000
assert cierre["efectivo_esperado"] == 57000
assert cierre["efectivo_contado"] == 57000
assert cierre["diferencia"] == 0
assert cierre["usuario"] == "admin"
print("PRUEBA A3b: OK -- el cierre queda guardado en cierres_caja con los valores correctos")

# A4. El "efectivo_esperado" SIEMPRE se recalcula en el servidor -- no
#     existe ni un campo de formulario para mandarlo manipulado, así
#     que probamos que, aunque cambiemos solo el contado, el esperado
#     guardado siga siendo base+ventas (no lo que hubiera puesto el
#     navegador).
_, tk = get_con_token("/caja/cerrar")
r = client.post("/caja/cerrar", data={
    "fecha": hoy, "base_inicial": "50000", "efectivo_contado": "60000", "observaciones": "sobran 3000",
    "csrf_token": tk,
}, follow_redirects=True)
assert r.status_code == 200
conn = appmod.get_db()
cierre = conn.execute("SELECT * FROM cierres_caja WHERE fecha=?", (hoy,)).fetchone()
conn.close()
assert cierre["efectivo_esperado"] == 57000, "FALLO: el esperado debe recalcularse siempre del lado del servidor"
assert cierre["efectivo_contado"] == 60000
assert cierre["diferencia"] == 3000
print("PRUEBA A4: OK -- el efectivo esperado se recalcula del lado del servidor, nunca se confía en el cliente")

# A5. El re-cierre queda en auditoría como "recerrar_caja", y el cierre
#     original como "cerrar_caja".
conn = appmod.get_db()
acciones = [row["accion"] for row in conn.execute(
    "SELECT accion FROM auditoria WHERE tabla_afectada='cierres_caja' ORDER BY id"
).fetchall()]
conn.close()
assert acciones == ["cerrar_caja", "recerrar_caja"], f"FALLO: acciones inesperadas en auditoría: {acciones}"
print("PRUEBA A5: OK -- el primer cierre y el re-cierre quedan diferenciados en auditoría")

# A6. La página /caja muestra el cierre reciente en su historial.
r = client.get("/caja")
assert r.status_code == 200
html = r.get_data(as_text=True)
assert "Cuadres de Caja" in html
assert hoy in html
print("PRUEBA A6: OK -- /caja muestra el historial de cuadres recientes")

# ═════════════════════════════════════════════════════════════════
# PRUEBA B: PDF DE TIQUETES DEL DÍA
# ═════════════════════════════════════════════════════════════════

limpiar_db()
set_sesion("admin", "admin")
hoy_dt = appmod.get_colombia_time()
hoy = hoy_dt.strftime("%Y-%m-%d")
ayer = (hoy_dt - appmod.timedelta(days=1)).strftime("%Y-%m-%d")

# B1. Un día sin ningún movimiento -> redirige de vuelta a /caja con aviso.
r = client.get(f"/tickets_dia_pdf?fecha=2000-01-01", follow_redirects=False)
assert r.status_code == 302 and "/caja" in r.headers["Location"]
print("PRUEBA B1: OK -- un día sin tiquetes no genera un PDF vacío, redirige con aviso")

# B2. Vehículo que solo entró hoy (sigue adentro) -> 1 tiquete (ENTRADA)
crear_entrada_salida("PDF001", "carro", 0, fecha_entrada=f"{hoy} 10:00:00", fecha_salida=False, marca="Renault")

# B3. Vehículo que entró y salió hoy -> 2 tiquetes (ENTRADA + SALIDA)
crear_entrada_salida("PDF002", "moto", 7000, fecha_entrada=f"{hoy} 11:00:00", fecha_salida=f"{hoy} 12:00:00", marca="AKT")

# B4. Vehículo que entró AYER y salió HOY (parqueo nocturno) -> en el
#     PDF de HOY solo debe salir el tiquete de SALIDA (la entrada ya
#     salió en el PDF de ayer).
crear_entrada_salida("PDF003", "carro", 30000, fecha_entrada=f"{ayer} 20:00:00", fecha_salida=f"{hoy} 08:00:00", marca="Chevrolet")

r = client.get(f"/tickets_dia_pdf?fecha={hoy}")
assert r.status_code == 200
assert r.headers["Content-Type"] == "application/pdf"
contenido = r.get_data()
assert len(contenido) > 500
print(f"PRUEBA B2: OK -- el PDF del día se genera correctamente ({len(contenido)} bytes)")

if TIENE_PDFPLUMBER:
    texto = texto_pdf(contenido)
    assert texto.count("PARQUEADERO EL PUENTE") == 4, (
        f"FALLO: se esperaban 4 tiquetes (PDF001-entrada, PDF002-entrada, PDF002-salida, PDF003-salida), "
        f"se encontraron {texto.count('PARQUEADERO EL PUENTE')}"
    )
    assert "PDF001" in texto and "ENTRADA" in texto
    assert "PDF002" in texto and "SALIDA" in texto
    assert "PDF003" in texto
    # PDF003 entró ayer -- su tiquete de ENTRADA NO debe estar en el PDF de hoy.
    idx_pdf003 = texto.find("PDF003")
    fragmento_pdf003 = texto[max(0, idx_pdf003 - 200):idx_pdf003]
    assert "SALIDA" in texto[idx_pdf003:idx_pdf003 + 50] or "Ticket:" in fragmento_pdf003
    print("PRUEBA B3: OK -- el PDF trae exactamente los tiquetes esperados, agrupados por vehículo")

    # El tiquete de entrada de PDF003 debe estar en el PDF de AYER, no en el de hoy.
    r_ayer = client.get(f"/tickets_dia_pdf?fecha={ayer}")
    texto_ayer = texto_pdf(r_ayer.get_data())
    assert "PDF003" in texto_ayer and "ENTRADA" in texto_ayer
    assert "PDF001" not in texto_ayer and "PDF002" not in texto_ayer
    print("PRUEBA B4: OK -- un vehículo con entrada ayer y salida hoy separa sus tiquetes en el PDF de cada día")
else:
    print("PRUEBA B3/B4: OMITIDA -- pdfplumber no está instalado en este entorno")

# B5. Un pago anulado se marca "PAGO ANULADO" SOLO en el tiquete de
#     salida (el de entrada es previo a cualquier cobro).
conn = appmod.get_db()
id_pdf002 = conn.execute("SELECT id FROM parqueadero WHERE placa='PDF002'").fetchone()["id"]
conn.execute("UPDATE parqueadero SET anulado=1, motivo_anulacion='prueba' WHERE id=?", (id_pdf002,))
conn.commit()
conn.close()

r = client.get(f"/tickets_dia_pdf?fecha={hoy}")
if TIENE_PDFPLUMBER:
    texto = texto_pdf(r.get_data())
    assert texto.count("PAGO ANULADO") == 1, f"FALLO: se esperaba 1 marca de anulado, hubo {texto.count('PAGO ANULADO')}"
    idx_entrada_pdf002 = texto.find("PDF002")
    # La primera aparición de PDF002 es su tiquete de ENTRADA -- ahí NO debe decir anulado.
    bloque_entrada = texto[idx_entrada_pdf002:idx_entrada_pdf002 + 250]
    assert "PAGO ANULADO" not in bloque_entrada, "FALLO: el tiquete de ENTRADA no debería marcarse como anulado"
    print("PRUEBA B5: OK -- 'PAGO ANULADO' aparece solo en el tiquete de SALIDA, nunca en el de ENTRADA")
else:
    print("PRUEBA B5: OMITIDA -- pdfplumber no está instalado en este entorno")

# ═════════════════════════════════════════════════════════════════
# PRUEBA C: EXCEL DE CLIENTES (CRM)
# ═════════════════════════════════════════════════════════════════

limpiar_db()
set_sesion("admin", "admin")

conn = appmod.get_db()
# Cliente de mensualidad, con teléfono y vencimiento.
conn.execute("""
    INSERT INTO mensualidades (nombre, placa, tipo, estado, fecha_inicio, fecha_fin, telefono)
    VALUES ('Juan Perez', 'MEN001', 'carro', 'Activo', '2026-01-01', '2026-12-31', '3001112222')
""")
# Cliente frecuente (diario) que NO tiene mensualidad.
conn.execute("INSERT INTO clientes_frecuentes (placa, marca, celular) VALUES ('DIA001', 'Mazda', '3003334444')")
# Placa que aparece en AMBAS tablas (por ejemplo, fue diario antes de
# volverse mensual) -- no debe salir duplicada, debe primar como Mensualidad.
conn.execute("INSERT INTO clientes_frecuentes (placa, marca, celular) VALUES ('MEN001', 'Toyota', '3005556666')")
conn.commit()
conn.close()

r = client.get("/mensualidades/excel_clientes")
assert r.status_code == 200
assert r.headers["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

wb = load_workbook(io.BytesIO(r.get_data()))
ws = wb.active
filas = list(ws.iter_rows(values_only=True))
assert filas[0] == ("Placa", "Telefono", "Tipo", "Vencimiento")
print("PRUEBA C1: OK -- el Excel trae exactamente las columnas Placa, Telefono, Tipo, Vencimiento")

datos = {fila[0]: fila for fila in filas[1:]}
assert datos["MEN001"][1] == "3001112222"
assert datos["MEN001"][2] == "Mensualidad"
assert datos["MEN001"][3] == "2026-12-31"
print("PRUEBA C2: OK -- el cliente de mensualidad sale con su teléfono y fecha de vencimiento")

assert datos["DIA001"][2] == "Diario"
assert datos["DIA001"][3] in (None, "")
print("PRUEBA C3: OK -- el cliente diario sale sin fecha de vencimiento")

assert len([f for f in filas[1:] if f[0] == "MEN001"]) == 1, (
    "FALLO: una placa que está en mensualidades Y en clientes_frecuentes no debe salir duplicada"
)
print("PRUEBA C4: OK -- una placa que es mensualidad no aparece duplicada como cliente diario")

# C5. Un operador (no-admin) no puede descargar la base de clientes.
set_sesion("operador1", "operador")
r = client.get("/mensualidades/excel_clientes", follow_redirects=False)
assert r.status_code == 302, "FALLO: un operador no debería poder descargar la base de clientes (datos sensibles)"
print("PRUEBA C5: OK -- un operador (no-admin) no puede descargar la base de clientes")

print()
print("TODAS LAS PRUEBAS DE CUADRE DE CAJA, PDF Y CRM PASARON")
