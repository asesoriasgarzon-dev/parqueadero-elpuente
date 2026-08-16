from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, flash
import sqlite3, os, math, shutil, secrets, hmac
import pytz
import calendar
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash, check_password_hash
from flask_wtf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

app = Flask(__name__)

# ─── SECRET_KEY ──────────────────────────────────────────────────
# Debe venir de una variable de entorno en producción (Railway).
# Si no está definida, se genera una temporal SOLO para que la app
# siga funcionando en desarrollo local — las sesiones se invalidan
# en cada reinicio, así que en producción SIEMPRE hay que definirla.
SECRET_KEY = os.environ.get("SECRET_KEY")
if not SECRET_KEY:
    SECRET_KEY = secrets.token_hex(32)
    print("ADVERTENCIA: SECRET_KEY no está definida en variables de entorno. "
          "Se generó una clave temporal solo para esta ejecución. "
          "Configura SECRET_KEY en producción (Railway).")
app.secret_key = SECRET_KEY

# ─── Cookies de sesión más seguras ────────────────────────────────
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
# En Railway (donde existe /data) se sirve por HTTPS -> cookie "Secure".
# En desarrollo local (HTTP) se deja en False para no romper el login.
app.config["SESSION_COOKIE_SECURE"] = os.path.exists("/data")

# ─── Protección CSRF global para todos los formularios POST ──────
csrf = CSRFProtect(app)

# ─── Límite de intentos de login (mitiga fuerza bruta) ────────────
limiter = Limiter(get_remote_address, app=app, default_limits=[])

POLIZA_NUM = "C-250004843"
CERTIFICADO_NUM = "10402089"
VIGENCIA_POLIZA = "Marzo 14 de 2026 a Marzo 14 de 2027"

# ─── Clave del agente de impresión ─────────────────────────────────
# El agente que corre en el PC de la caseta usa esta misma clave
# para poder consultar/marcar los tickets pendientes de imprimir.
# Ya NO tiene un valor por defecto: si no está configurada, la API
# de impresión queda deshabilitada (no expuesta con una clave conocida).
PRINTER_API_KEY = os.environ.get("PRINTER_API_KEY")
if not PRINTER_API_KEY:
    print("ADVERTENCIA: PRINTER_API_KEY no está definida. "
          "La impresión remota (agente local) quedará deshabilitada hasta que la configures.")

def get_colombia_time():
    # Creamos la zona horaria de Bogotá
    tz = pytz.timezone('America/Bogota')
    # Devolvemos la hora actual en esa zona
    return datetime.now(tz)

# ─── CONFIGURACIÓN DE RUTA PARA RAILWAY VOLUME ───
# Si la carpeta /data existe (en Railway), usamos esa ruta. 
# Si no (en tu PC), usa la ruta local.
if os.path.exists('/data'):
    DB_FILE = "/data/parqueadero.db"
    BACKUP_DIR = "/data/backups"
else:
    DB_FILE = "parqueadero.db"
    BACKUP_DIR = "backups"

if not os.path.exists(BACKUP_DIR):
    os.makedirs(BACKUP_DIR)

# ─── Base de datos ────────────────────────────────────────────────
def get_db():
    # timeout=10: si otra conexión tiene el archivo bloqueado escribiendo,
    # espera hasta 10s reintentando en vez de fallar de inmediato con
    # "database is locked" (fundamental con varios operadores a la vez).
    conn = sqlite3.connect(DB_FILE, timeout=10)
    conn.row_factory = sqlite3.Row
    # WAL: permite que las lecturas (consultas, listados) no se bloqueen
    # mientras otra conexión está escribiendo una entrada/salida.
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=8000")
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    
    # 1. Tabla Parqueadero
    c.execute("""CREATE TABLE IF NOT EXISTS parqueadero (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        placa TEXT, 
        tipo TEXT,
        hora_entrada TEXT, 
        hora_salida TEXT,
        valor REAL, 
        ticket_num INTEGER
    )""")
    
    # 2. Tabla Consecutivo
    c.execute("""CREATE TABLE IF NOT EXISTS consecutivo (
        id INTEGER PRIMARY KEY, numero INTEGER
    )""")
    c.execute("INSERT OR IGNORE INTO consecutivo (id, numero) VALUES (1, 0)")
    
    # 3. Tabla Mensualidades
    c.execute("""CREATE TABLE IF NOT EXISTS mensualidades (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT, 
        placa TEXT UNIQUE,
        tipo TEXT, 
        estado TEXT,
        fecha_inicio TEXT, 
        fecha_fin TEXT
    )""")
    
    # 4. Tabla Clientes Frecuentes
    c.execute("""CREATE TABLE IF NOT EXISTS clientes_frecuentes (
        placa TEXT PRIMARY KEY, marca TEXT, celular TEXT
    )""")
    
    # 5. Tabla Usuarios
    c.execute("""CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE,
        password TEXT,
        rol TEXT DEFAULT 'operador'
    )""")
    
    # 6. Tabla Tarifas
    c.execute("""CREATE TABLE IF NOT EXISTS tarifas (
        id INTEGER PRIMARY KEY,
        tipo TEXT UNIQUE,
        valor_hora REAL,
        valor_dia REAL,
        minutos_cortesia INTEGER DEFAULT 5
    )""")

    # 7. Tabla Auditoría (trazabilidad de cambios sensibles)
    # Registra quién hizo qué, cuándo, sobre qué registro, y los valores
    # antes/después. No reemplaza el histórico de `parqueadero`, lo
    # complementa para acciones que MODIFICAN o BORRAN algo ya existente
    # (anulaciones, cambios de tarifa, reset de clave, borrado de
    # mensualidades) -- ahí es donde antes no quedaba ningún rastro.
    c.execute("""CREATE TABLE IF NOT EXISTS auditoria (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha_hora TEXT,
        usuario TEXT,
        accion TEXT,
        tabla_afectada TEXT,
        registro_id TEXT,
        valor_anterior TEXT,
        valor_nuevo TEXT,
        motivo TEXT
    )""")

    # Datos iniciales (Usuarios y Tarifas 2026)
    # NOTA: en instalaciones ya existentes, INSERT OR IGNORE no toca las filas
    # que ya están ahí, así que esto no afecta contraseñas ya migradas a hash.
    c.execute("INSERT OR IGNORE INTO usuarios (username, password, rol) VALUES ('admin', ?, 'admin')",
              (generate_password_hash("admin123"),))
    c.execute("INSERT OR IGNORE INTO usuarios (username, password, rol) VALUES ('operador', ?, 'operador')",
              (generate_password_hash("op123"),))
    c.execute("INSERT OR IGNORE INTO tarifas (id, tipo, valor_hora, valor_dia, minutos_cortesia) VALUES (1, 'carro', 4000, 14000, 5)")
    c.execute("INSERT OR IGNORE INTO tarifas (id, tipo, valor_hora, valor_dia, minutos_cortesia) VALUES (2, 'moto', 2500, 7000, 5)")
    
    # --- BLOQUE DE ACTUALIZACIÓN SEGURO (Evolución de la BD) ---
    
    # Columnas adicionales para Parqueadero (Control de auditoría y anulaciones)
    columnas_parqueadero = [
        ("marca", "TEXT"),
        ("celular", "TEXT"),
        ("metodo_pago", "TEXT DEFAULT 'Efectivo'"),
        ("cajero", "TEXT DEFAULT 'Operador'"),
        ("observaciones", "TEXT"),
        ("anulado", "INTEGER DEFAULT 0"),
        ("motivo_anulacion", "TEXT"),
        ("valor_real", "REAL DEFAULT 0"),
        ("impreso_entrada", "INTEGER DEFAULT 0"),
        ("impreso_salida", "INTEGER DEFAULT 0")
    ]
    
    for col in columnas_parqueadero:
        try:
            c.execute(f"ALTER TABLE parqueadero ADD COLUMN {col[0]} {col[1]}")
        except Exception:
            pass # La columna ya existe

    # Columnas adicionales para Mensualidades (Datos del propietario)
    columnas_mensualidades = [
        ("telefono", "TEXT"), 
        ("modelo", "TEXT"), 
        ("color", "TEXT"), 
        ("observaciones", "TEXT")
    ]
    
    for col in columnas_mensualidades:
        try:
            c.execute(f"ALTER TABLE mensualidades ADD COLUMN {col[0]} {col[1]}")
        except Exception:
            pass # La columna ya existe
    
    # Actualizar tarifa diaria del carro a $14.000
    c.execute("""
        UPDATE tarifas
        SET valor_dia = 14000
        WHERE tipo = 'carro'
    """)
    
    conn.commit()
    conn.close()
    print("Base de datos inicializada y actualizada correctamente.")
# ─── Auth ────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "usuario" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("rol") != "admin":
            return redirect(url_for("inicio"))
        return f(*args, **kwargs)
    return decorated

# ─── Helpers ─────────────────────────────────────────────────────
def get_tarifas():
    conn = get_db()
    regs = conn.execute("SELECT tipo, valor_hora, valor_dia, minutos_cortesia FROM tarifas").fetchall()
    conn.close()
    # Convertimos a un diccionario para que sea fácil de usar: {'carro': {...}, 'moto': {...}}
    return {r["tipo"]: dict(r) for r in regs}

def calcular_valor(tipo, minutos_totales, tarifas):
    t = tarifas[tipo]
    cortesia = t["minutos_cortesia"]

    # Cortesía inicial
    if minutos_totales <= cortesia:
        return 0

    if tipo == "carro":

        # Primera hora o fracción
        if minutos_totales <= 60:
            total = t["valor_hora"]

        else:
            # Horas/fracciones adicionales
            fracciones = math.ceil((minutos_totales - 60) / 60)

            # Primera hora $4.000
            # Cada fracción adicional $3.500
            total = t["valor_hora"] + fracciones * (t["valor_hora"] - 500)

        # Tope máximo diario: $14.000
        return min(total, t["valor_dia"])

    else:

        # Moto: $2.500 por hora o fracción
        total = math.ceil(minutos_totales / 60) * t["valor_hora"]

        # Tope máximo diario: $7.000
        return min(total, t["valor_dia"])

def get_consecutivo(conn):
    """
    Entrega el siguiente número de ticket (1, 2, 3...).

    IMPORTANTE: a propósito NO se hace conn.commit() aquí. En SQLite, el
    UPDATE de abajo toma el bloqueo de escritura de la base de datos desde
    el momento en que se ejecuta, y lo mantiene hasta que la conexión que
    llamó a esta función haga su propio commit() más adelante (después de
    insertar la fila en `parqueadero`). Mientras ese commit no ocurra,
    ninguna otra conexión puede tomar un número de ticket, así que dos
    operadores registrando una entrada al mismo tiempo NUNCA pueden recibir
    el mismo consecutivo. Si esta función hiciera su propio commit() antes
    de que el llamador termine, se abriría una ventana en la que otra
    conexión podría colarse y repetir el número (el bug original).
    """
    c = conn.cursor()
    c.execute("UPDATE consecutivo SET numero = numero + 1 WHERE id = 1")
    res = conn.execute("SELECT numero FROM consecutivo WHERE id = 1").fetchone()
    return res["numero"]

def registrar_auditoria(conn, accion, tabla_afectada, registro_id,
                         valor_anterior=None, valor_nuevo=None, motivo=None):
    """
    Deja un rastro permanente de una acción sensible: quién, qué, cuándo,
    sobre qué registro, y los valores antes/después.

    Se inserta con el mismo `conn` (misma conexión/transacción) que el
    cambio que audita, y el llamador es quien hace el commit() -- así, si
    el cambio se cancela (rollback), el registro de auditoría se cancela
    con él y nunca queda un rastro de algo que en realidad no ocurrió.

    Nunca se le pasa contraseñas ni datos de tarjetas/pagos sensibles a
    valor_anterior/valor_nuevo -- solo lo mínimo para poder reconstruir
    qué cambió.
    """
    conn.execute("""
        INSERT INTO auditoria
            (fecha_hora, usuario, accion, tabla_afectada, registro_id, valor_anterior, valor_nuevo, motivo)
        VALUES (?,?,?,?,?,?,?,?)
    """, (
        get_colombia_time().strftime("%Y-%m-%d %H:%M:%S"),
        session.get("usuario", "sistema"),
        accion,
        tabla_afectada,
        str(registro_id) if registro_id is not None else None,
        str(valor_anterior) if valor_anterior is not None else None,
        str(valor_nuevo) if valor_nuevo is not None else None,
        motivo,
    ))

def _normalizar_fecha_excel(valor):
    """Convierte un valor de celda de Excel (fecha o texto) a 'AAAA-MM-DD'."""
    if valor is None or valor == "":
        return ""
    if hasattr(valor, "strftime"):  # datetime.datetime o datetime.date
        return valor.strftime("%Y-%m-%d")
    texto = str(valor).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(texto, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return texto  # se guarda tal cual si no se pudo interpretar

def _es_hash(valor):
    """Detecta si un password guardado ya es un hash de werkzeug (pbkdf2/scrypt)."""
    return bool(valor) and (valor.startswith("pbkdf2:") or valor.startswith("scrypt:"))

def _password_valida(password_guardado, password_ingresada):
    """
    Compara la contraseña ingresada contra la guardada.
    Soporta migración transparente: si el usuario todavía tiene la
    contraseña en texto plano (instalaciones antiguas), se compara
    directo; si ya es un hash, se valida con check_password_hash.
    """
    if _es_hash(password_guardado):
        return check_password_hash(password_guardado, password_ingresada)
    return password_guardado == password_ingresada

# ─── Rutas ───────────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
@limiter.limit("8 per minute")
def login():
    error = None
    if request.method == "POST":
        u = request.form.get("username", "").strip()
        p = request.form.get("password", "").strip()
        conn = get_db()
        user = conn.execute("SELECT * FROM usuarios WHERE username=?", (u,)).fetchone()

        if user and _password_valida(user["password"], p):
            # Migración transparente: si todavía estaba en texto plano,
            # se re-guarda como hash ahora que sabemos que es correcta.
            if not _es_hash(user["password"]):
                conn.execute("UPDATE usuarios SET password=? WHERE id=?",
                             (generate_password_hash(p), user["id"]))
                conn.commit()
            session["usuario"] = user["username"]
            session["rol"] = user["rol"]
            conn.close()
            return redirect(url_for("inicio"))

        conn.close()
        error = "Usuario o contraseña incorrectos"
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/")
@login_required
def inicio():
    conn = get_db()
    activos = conn.execute("SELECT COUNT(*) as c FROM parqueadero WHERE hora_salida IS NULL").fetchone()["c"]
   # CAMBIO AQUÍ: Usamos get_colombia_time() en lugar de datetime.now()
    ahora_co = get_colombia_time() 
    hoy = ahora_co.strftime("%Y-%m-%d")
    caja_hoy = conn.execute("""
        SELECT COALESCE(SUM(valor),0) as t FROM parqueadero
        WHERE date(hora_salida)=? AND (anulado = 0 OR anulado IS NULL)
    """, (hoy,)).fetchone()["t"]
    conn.close()
    return render_template("inicio.html", activos=activos, caja_hoy=int(caja_hoy))

# ─── Entrada ─────────────────────────────────────────────────────
@app.route("/entrada/<tipo>", methods=["GET","POST"])
@login_required
def entrada(tipo):
    if tipo not in ("carro", "moto"):
        return "Tipo de vehículo no válido", 404
    mensaje = None
    cliente = None
    if request.method == "POST":
        placa = request.form.get("placa","").strip().upper()
        marca = request.form.get("marca","").strip()
        celular = request.form.get("celular","").strip()
        obs = request.form.get("observaciones","").strip()

        if not placa:
            mensaje = ("error", "La placa es obligatoria.")
        else:
            conn = get_db()
            
            # --- LÍNEA SALVAVIDAS PARA EVITAR ERROR 500 ---
            try:
                conn.execute("ALTER TABLE parqueadero ADD COLUMN observaciones TEXT")
            except:
                pass # Si la columna ya existe, no hace nada
            # ----------------------------------------------

            existe = conn.execute("SELECT id FROM parqueadero WHERE placa=? AND hora_salida IS NULL", (placa,)).fetchone()
            if existe:
                conn.close()
                return redirect(url_for("salida_form", placa=placa))
            
            ticket_num = get_consecutivo(conn)
            ahora = get_colombia_time().strftime("%Y-%m-%d %H:%M:%S")
            
            # INSERT con los 8 campos
            conn.execute("""INSERT INTO parqueadero 
                (placa, tipo, hora_entrada, ticket_num, marca, celular, cajero, observaciones) 
                VALUES (?,?,?,?,?,?,?,?)""",
                (placa, tipo, ahora, ticket_num, marca, celular, session.get("usuario"), obs))
            
            conn.execute("INSERT OR REPLACE INTO clientes_frecuentes (placa, marca, celular) VALUES (?,?,?)",
                         (placa, marca, celular))
            conn.commit()
            conn.close()
            return redirect(url_for("ticket_entrada", ticket=ticket_num))
    
    else:
        placa_pre = request.args.get("placa","")
        if placa_pre:
            conn = get_db()
            cliente = conn.execute("SELECT * FROM clientes_frecuentes WHERE placa=?", (placa_pre,)).fetchone()
            conn.close()
    return render_template("entrada.html", tipo=tipo, mensaje=mensaje, cliente=cliente,
                           placa_pre=request.args.get("placa",""))

@app.route("/buscar_placa")
@login_required
def buscar_placa():
    placa = request.args.get("placa","").strip().upper()
    if not placa:
        return jsonify({"status": "vacio"})
    conn = get_db()

    # 1. Verificar si ya está en el patio
    activo = conn.execute("SELECT id FROM parqueadero WHERE placa=? AND hora_salida IS NULL", (placa,)).fetchone()
    if activo:
        conn.close()
        return jsonify({"status": "activo", "placa": placa})

    # 2. Verificar mensualidad
    hoy = get_colombia_time().strftime("%Y-%m-%d")
    mensualidad = conn.execute("""
        SELECT nombre, fecha_fin, estado FROM mensualidades
        WHERE placa=? ORDER BY fecha_fin DESC LIMIT 1
    """, (placa,)).fetchone()

    mensualidad_info = None
    if mensualidad:
        if mensualidad["estado"] == "Activo" and mensualidad["fecha_fin"] >= hoy:
            mensualidad_info = {"tipo": "activa", "hasta": mensualidad["fecha_fin"], "nombre": mensualidad["nombre"]}
        else:
            mensualidad_info = {"tipo": "vencida", "hasta": mensualidad["fecha_fin"], "nombre": mensualidad["nombre"]}

    # 3. Verificar cliente frecuente
    cliente = conn.execute("SELECT * FROM clientes_frecuentes WHERE placa=?", (placa,)).fetchone()
    conn.close()

    if cliente:
        return jsonify({"status": "conocido", "marca": cliente["marca"], "celular": cliente["celular"], "mensualidad": mensualidad_info})
    return jsonify({"status": "nuevo", "mensualidad": mensualidad_info})

# ─── Salida ──────────────────────────────────────────────────────
@app.route("/salida", methods=["GET", "POST"])
@login_required
def salida_form():

    # ============================================================
    # VARIABLES INICIALES
    # ============================================================

    placa = request.args.get("placa", "").strip().upper()

    reg = None
    valor_sugerido = 0
    tiempo_str = ""

    # ============================================================
    # CONSULTA DE VEHÍCULO PARA MOSTRAR LA SALIDA
    # ============================================================

    if placa:

        conn = get_db()

        reg = conn.execute("""
            SELECT *
            FROM parqueadero
            WHERE placa = ?
              AND hora_salida IS NULL
            ORDER BY id DESC
            LIMIT 1
        """, (placa,)).fetchone()

        conn.close()

        if reg:

            # ----------------------------------------------------
            # Convertir entrada a hora Colombia
            # ----------------------------------------------------

            entrada_dt = datetime.strptime(
                reg["hora_entrada"],
                "%Y-%m-%d %H:%M:%S"
            )

            tz = pytz.timezone("America/Bogota")

            entrada_dt = tz.localize(entrada_dt)

            # ----------------------------------------------------
            # Hora actual Colombia
            # ----------------------------------------------------

            ahora_co = get_colombia_time()

            # ----------------------------------------------------
            # Calcular duración real
            # ----------------------------------------------------

            dur = ahora_co - entrada_dt

            mins = int(
                max(
                    0,
                    dur.total_seconds() // 60
                )
            )

            h, m = divmod(mins, 60)

            tiempo_str = f"{h}h {m}m"

            # ----------------------------------------------------
            # Calcular valor según tarifas vigentes
            # ----------------------------------------------------

            tarifas = get_tarifas()

            valor_sugerido = int(
                calcular_valor(
                    reg["tipo"],
                    mins,
                    tarifas
                )
            )


    # ============================================================
    # REGISTRAR SALIDA
    # ============================================================

    if request.method == "POST":

        placa = request.form.get(
            "placa",
            ""
        ).strip().upper()

        valor_raw = request.form.get(
            "valor",
            "0"
        ).strip()

        metodo = request.form.get(
            "metodo_pago",
            "Efectivo"
        ).strip()

        convenio = request.form.get(
            "convenio",
            ""
        ).strip()


        # --------------------------------------------------------
        # Convertir valor recibido a número
        # --------------------------------------------------------

        try:

            valor_limpio = (
                str(valor_raw)
                .replace("$", "")
                .replace(".", "")
                .replace(",", "")
                .strip()
            )

            valor_pago = int(
                float(valor_limpio)
            )

        except (ValueError, TypeError):

            valor_pago = 0


        # --------------------------------------------------------
        # Buscar nuevamente el vehículo
        #
        # IMPORTANTE:
        # No usamos el registro obtenido en GET.
        # Volvemos a consultar la BD para asegurarnos de que
        # el vehículo todavía esté dentro del parqueadero.
        # --------------------------------------------------------

        conn = get_db()

        reg2 = conn.execute("""
            SELECT *
            FROM parqueadero
            WHERE placa = ?
              AND hora_salida IS NULL
            ORDER BY id DESC
            LIMIT 1
        """, (placa,)).fetchone()


        if reg2:

            # ----------------------------------------------------
            # Hora oficial de salida Colombia
            # ----------------------------------------------------

            ahora = get_colombia_time().strftime(
                "%Y-%m-%d %H:%M:%S"
            )


            # ----------------------------------------------------
            # Construir método de pago + convenio
            # ----------------------------------------------------

            metodo_pago = metodo

            if convenio:

                metodo_pago = (
                    f"{metodo} - {convenio}"
                )


            # ----------------------------------------------------
            # Registrar salida
            #
            # valor = valor REAL cobrado al cliente.
            #
            # No recalculamos aquí el valor porque el cajero
            # puede haber seleccionado:
            #
            # - Valor por tiempo
            # - Tarifa de día
            # - Un valor autorizado/manual
            # ----------------------------------------------------
            #
            # NOTA sobre concurrencia (doble salida):
            #
            # Entre el SELECT de arriba y este UPDATE, otro operador
            # pudo haber registrado la salida de este mismo vehículo
            # (dos cajeros atendiendo el mismo carro casi al mismo
            # tiempo). Por eso el UPDATE repite la condición
            # "hora_salida IS NULL": si alguien más ya lo cerró, esta
            # sentencia no actualiza ninguna fila (rowcount = 0) y lo
            # detectamos abajo, en vez de sobreescribir en silencio
            # el cobro/ticket que el otro operador ya generó.
            # ----------------------------------------------------

            cur = conn.execute("""
                UPDATE parqueadero
                SET
                    hora_salida = ?,
                    valor = ?,
                    metodo_pago = ?,
                    cajero = ?
                WHERE id = ?
                  AND hora_salida IS NULL
            """, (
                ahora,
                valor_pago,
                metodo_pago,
                session.get("usuario"),
                reg2["id"]
            ))

            if cur.rowcount == 0:
                # Alguien más ganó la carrera: ya se registró la salida
                # de este vehículo entre que lo consultamos y confirmamos.
                conn.rollback()
                conn.close()
                flash(
                    "Este vehículo ya fue despachado por otro operador "
                    "justo antes de confirmar. Verifica el histórico.",
                    "error"
                )
                return redirect(url_for("salida_form"))

            conn.commit()

            ticket = reg2["ticket_num"]

            conn.close()


            # ----------------------------------------------------
            # Generar ticket de salida
            # ----------------------------------------------------

            return redirect(
                url_for(
                    "ticket_salida",
                    ticket=ticket
                )
            )


        # --------------------------------------------------------
        # Si el vehículo ya no está disponible
        # (ya salió, o la placa no existe / no está en el patio)
        # --------------------------------------------------------

        conn.close()
        flash(
            "No se encontró ese vehículo activo en el patio "
            "(puede que ya haya salido).",
            "error"
        )
        return redirect(url_for("salida_form"))


    # ============================================================
    # MOSTRAR FORMULARIO
    # ============================================================

    return render_template(
        "salida.html",
        placa=placa,
        reg=reg,
        valor_sugerido=valor_sugerido,
        tiempo_str=tiempo_str
    )

# ─── Tickets (vista imprimible) ───────────────────────────────────
@app.route("/ticket/entrada/<int:ticket>")
@login_required
def ticket_entrada(ticket):
    conn = get_db()
    reg = conn.execute("SELECT * FROM parqueadero WHERE ticket_num=?", (ticket,)).fetchone()
    conn.close()
    tarifas = get_tarifas()
    # PASAMOS LOS DATOS DE LA PÓLIZA AQUÍ
    return render_template("ticket.html", 
                           reg=reg, 
                           tipo="entrada", 
                           tarifas=tarifas,
                           poliza=POLIZA_NUM, 
                           certificado=CERTIFICADO_NUM, 
                           vigencia=VIGENCIA_POLIZA)

@app.route("/ticket/salida/<int:ticket>")
@login_required
def ticket_salida(ticket):
    conn = get_db()
    reg = conn.execute("SELECT * FROM parqueadero WHERE ticket_num=?", (ticket,)).fetchone()
    conn.close()
    
    if reg and reg["hora_entrada"] and reg["hora_salida"]:
        entrada_dt = datetime.strptime(reg["hora_entrada"], "%Y-%m-%d %H:%M:%S")
        salida_dt = datetime.strptime(reg["hora_salida"], "%Y-%m-%d %H:%M:%S")
        dur = salida_dt - entrada_dt
        mins = int(max(0, dur.total_seconds() // 60))
        h, m = divmod(mins, 60)
        tiempo_str = f"{h}h {m}m"
    else:
        tiempo_str = ""
    
    # TAMBIÉN LOS PASAMOS EN LA SALIDA
    return render_template("ticket.html", 
                           reg=reg, 
                           tipo="salida", 
                           tiempo_str=tiempo_str,
                           poliza=POLIZA_NUM, 
                           certificado=CERTIFICADO_NUM, 
                           vigencia=VIGENCIA_POLIZA)

# ─── Clientes activos ─────────────────────────────────────────────
@app.route("/clientes")
@login_required
def clientes():
    conn = get_db()
    activos = conn.execute("SELECT * FROM parqueadero WHERE hora_salida IS NULL ORDER BY hora_entrada").fetchall()
    conn.close()
    
    tarifas = get_tarifas()
    # CAMBIO CRÍTICO: Usamos la hora de Colombia para comparar con los carros que están adentro
    ahora_co = get_colombia_time()
    
    lista = []
    for r in activos:
        entrada_dt = datetime.strptime(r["hora_entrada"], "%Y-%m-%d %H:%M:%S")
        
        # Localizamos la hora de entrada para que sea compatible con la de Colombia
        tz = pytz.timezone('America/Bogota')
        entrada_dt = tz.localize(entrada_dt)
        
        # Restamos Colombia vs Colombia
        mins = int(max(0, (ahora_co - entrada_dt).total_seconds() // 60))
        
        h, m = divmod(mins, 60)
        valor_est = int(calcular_valor(r["tipo"], mins, tarifas))
        lista.append({"reg": r, "tiempo": f"{h}h {m}m", "valor_est": valor_est})
        
    return render_template("clientes.html", lista=lista)

# ─── Histórico ───────────────────────────────────────────────────
@app.route("/historico")
@login_required
def historico():
    placa = request.args.get("placa","").strip().upper()
    conn = get_db()
    if placa:
        regs = conn.execute("SELECT * FROM parqueadero WHERE placa LIKE ? ORDER BY id DESC LIMIT 100", (f"%{placa}%",)).fetchall()
    else:
        regs = conn.execute("SELECT * FROM parqueadero ORDER BY id DESC LIMIT 100").fetchall()
    conn.close()
    return render_template("historico.html", regs=regs, placa=placa)

# ─── Caja ────────────────────────────────────────────────────────
@app.route("/caja")
@login_required
def caja():
    ahora_co = get_colombia_time()
    fecha_hoy = ahora_co.strftime("%Y-%m-%d")
    
    # Recibimos ambos parámetros, si no existen, por defecto es hoy
    fecha_inicio = request.args.get("fecha_inicio", fecha_hoy)
    fecha_fin = request.args.get("fecha_fin", fecha_hoy)
    
    conn = get_db()
    # Cambiamos la consulta SQL para que use un rango (BETWEEN)
    # Se excluyen los pagos anulados: un cobro anulado no es un ingreso
    # real, no debe aparecer en el cierre de caja (sí sigue visible,
    # completo, en el Histórico -- ese es el lugar para auditarlos).
    query = """
        SELECT * FROM parqueadero
        WHERE date(hora_salida) BETWEEN ? AND ?
          AND (anulado = 0 OR anulado IS NULL)
        ORDER BY hora_salida DESC
    """
    regs = conn.execute(query, (fecha_inicio, fecha_fin)).fetchall()
    
    total = sum(r["valor"] for r in regs if r["valor"])
    por_metodo = {}
    for r in regs:
        m = r["metodo_pago"] or "Efectivo"
        metodo_base = m.split(" - ")[0]
        por_metodo[metodo_base] = por_metodo.get(metodo_base, 0) + (r["valor"] or 0)
    conn.close()
    
    return render_template("caja.html", regs=regs, total=int(total), 
                           fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, 
                           por_metodo=por_metodo)
# ─── Mensualidades ────────────────────────────────────────────────
@app.route("/mensualidades", methods=["GET", "POST"])
@login_required
def mensualidades():
    conn = get_db()
    
    # REPARACIÓN: Agrega las columnas de la dueña si no existen
    for col in ["telefono", "modelo", "color", "observaciones"]:
        try:
            conn.execute(f"ALTER TABLE mensualidades ADD COLUMN {col} TEXT")
            conn.commit()
        except:
            pass

    if request.method == "POST":
        # Crear/editar mensualidades queda restringido a admin (la
        # consulta/listado sigue abierta a cualquier operador logueado).
        if session.get("rol") != "admin":
            conn.close()
            return redirect(url_for("mensualidades"))

        nombre = request.form.get("nombre", "").strip()
        placa = request.form.get("placa", "").strip().upper()
        tel = request.form.get("telefono", "").strip()
        mod = request.form.get("modelo", "").strip()
        col_v = request.form.get("color", "").strip()
        tipo = request.form.get("tipo", "")
        fi = request.form.get("fecha_inicio", "")
        ff = request.form.get("fecha_fin", "")
        est = request.form.get("estado", "Activo")
        obs = request.form.get("observaciones", "").strip()

        check = conn.execute("SELECT id FROM mensualidades WHERE placa = ?", (placa,)).fetchone()

        if check:
            anterior = conn.execute("SELECT estado, fecha_fin FROM mensualidades WHERE id=?", (check["id"],)).fetchone()
            conn.execute("""UPDATE mensualidades SET
                nombre=?, telefono=?, modelo=?, color=?, tipo=?,
                fecha_inicio=?, fecha_fin=?, estado=?, observaciones=?
                WHERE placa=?""", (nombre, tel, mod, col_v, tipo, fi, ff, est, obs, placa))
            registrar_auditoria(
                conn,
                accion="editar_mensualidad",
                tabla_afectada="mensualidades",
                registro_id=check["id"],
                valor_anterior=f"estado={anterior['estado']}, fecha_fin={anterior['fecha_fin']}",
                valor_nuevo=f"estado={est}, fecha_fin={ff}",
                motivo=f"placa: {placa}",
            )
        else:
            cur = conn.execute("""INSERT INTO mensualidades
                (nombre, placa, telefono, modelo, color, tipo, fecha_inicio, fecha_fin, estado, observaciones)
                VALUES (?,?,?,?,?,?,?,?,?,?)""", (nombre, placa, tel, mod, col_v, tipo, fi, ff, est, obs))
            registrar_auditoria(
                conn,
                accion="crear_mensualidad",
                tabla_afectada="mensualidades",
                registro_id=cur.lastrowid,
                valor_nuevo=f"placa={placa}, nombre={nombre}, estado={est}",
                motivo=None,
            )
        conn.commit()
        return redirect(url_for('mensualidades'))

    regs = conn.execute("SELECT * FROM mensualidades ORDER BY nombre ASC").fetchall()
    hoy = get_colombia_time().date()
    lista = []
    for r in regs:
        fila = dict(r)
        alerta = "Activo"
        if fila.get('fecha_fin'):
            try:
                f_fin = datetime.strptime(fila['fecha_fin'], "%Y-%m-%d").date()
                if f_fin < hoy: alerta = "Vencida"
                elif f_fin <= hoy + timedelta(days=5): alerta = "Por vencer"
            except: pass
        lista.append({'reg': fila, 'alerta': alerta})
    conn.close()

    mensaje_carga = request.args.get("resultado")
    error_carga = request.args.get("error")

    return render_template("mensualidades.html", lista=lista,
                           mensaje_carga=mensaje_carga, error_carga=error_carga)

@app.route("/eliminar_mensualidad/<int:id>", methods=["POST"])
@login_required
@admin_required
def eliminar_mensualidad(id):
    conn = get_db()

    # Capturamos los datos antes de borrar -- una vez hecho el DELETE no
    # hay forma de recuperarlos, así que esta es la única oportunidad de
    # dejar un rastro de qué se eliminó.
    reg = conn.execute("SELECT nombre, placa, estado FROM mensualidades WHERE id=?", (id,)).fetchone()

    conn.execute("DELETE FROM mensualidades WHERE id = ?", (id,))

    if reg:
        registrar_auditoria(
            conn,
            accion="eliminar_mensualidad",
            tabla_afectada="mensualidades",
            registro_id=id,
            valor_anterior=f"nombre={reg['nombre']}, placa={reg['placa']}, estado={reg['estado']}",
            valor_nuevo="eliminado",
            motivo=None,
        )

    conn.commit()
    conn.close()
    return redirect(url_for('mensualidades'))

# ─── Plantilla y carga masiva de mensualidades por Excel ─────────
@app.route("/mensualidades/plantilla_excel")
@login_required
def plantilla_excel_mensualidades():
    wb = Workbook()
    ws = wb.active
    ws.title = "Mensualidades"

    encabezados = [
        "Nombre", "Placa", "Telefono", "Modelo", "Color",
        "Tipo (carro/moto)", "Fecha Inicio (AAAA-MM-DD)",
        "Fecha Fin (AAAA-MM-DD)", "Estado (Activo/Inactivo)", "Observaciones"
    ]
    ws.append(encabezados)
    ws.append([
        "Juan Perez", "ABC123", "3001234567", "Mazda 3", "Rojo",
        "carro", "2026-01-01", "2026-12-31", "Activo", "Cliente frecuente"
    ])

    for idx, encabezado in enumerate(encabezados, start=1):
        letra = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[letra].width = max(18, len(encabezado) + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="plantilla_mensualidades.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/mensualidades/cargar_excel", methods=["POST"])
@login_required
@admin_required
def cargar_excel_mensualidades():
    archivo = request.files.get("archivo_excel")
    if not archivo or archivo.filename == "":
        return redirect(url_for("mensualidades", error="No seleccionaste ningún archivo."))

    try:
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active
    except Exception as e:
        return redirect(url_for("mensualidades", error=f"No se pudo leer el archivo: {e}"))

    conn = get_db()
    insertados = 0
    actualizados = 0
    errores = []

    for i, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not fila or all(c is None for c in fila):
            continue
        try:
            nombre = str(fila[0]).strip() if len(fila) > 0 and fila[0] else ""
            placa = str(fila[1]).strip().upper() if len(fila) > 1 and fila[1] else ""
            telefono = str(fila[2]).strip() if len(fila) > 2 and fila[2] else ""
            modelo = str(fila[3]).strip() if len(fila) > 3 and fila[3] else ""
            color = str(fila[4]).strip() if len(fila) > 4 and fila[4] else ""
            tipo = str(fila[5]).strip().lower() if len(fila) > 5 and fila[5] else "carro"
            fecha_inicio = _normalizar_fecha_excel(fila[6]) if len(fila) > 6 else ""
            fecha_fin = _normalizar_fecha_excel(fila[7]) if len(fila) > 7 else ""
            estado = str(fila[8]).strip() if len(fila) > 8 and fila[8] else "Activo"
            observaciones = str(fila[9]).strip() if len(fila) > 9 and fila[9] else ""

            if not placa:
                errores.append(f"Fila {i}: falta la placa, se omitió")
                continue
            if tipo not in ("carro", "moto"):
                tipo = "carro"

            check = conn.execute("SELECT id FROM mensualidades WHERE placa = ?", (placa,)).fetchone()
            if check:
                conn.execute("""UPDATE mensualidades SET
                    nombre=?, telefono=?, modelo=?, color=?, tipo=?,
                    fecha_inicio=?, fecha_fin=?, estado=?, observaciones=?
                    WHERE placa=?""",
                    (nombre, telefono, modelo, color, tipo,
                     fecha_inicio, fecha_fin, estado, observaciones, placa))
                actualizados += 1
            else:
                conn.execute("""INSERT INTO mensualidades
                    (nombre, placa, telefono, modelo, color, tipo, fecha_inicio, fecha_fin, estado, observaciones)
                    VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (nombre, placa, telefono, modelo, color, tipo,
                     fecha_inicio, fecha_fin, estado, observaciones))
                insertados += 1
        except Exception as e:
            errores.append(f"Fila {i}: {e}")

    conn.commit()
    conn.close()

    mensaje = f"{insertados} mensualidad(es) nueva(s), {actualizados} actualizada(s)."
    if errores:
        mensaje += f" {len(errores)} fila(s) con problemas: " + " | ".join(errores[:5])

    return redirect(url_for("mensualidades", resultado=mensaje))
    
# ─── Tarifas (solo admin) ─────────────────────────────────────────
@app.route("/tarifas", methods=["GET","POST"])
@login_required
@admin_required
def tarifas():
    conn = get_db()
    if request.method == "POST":
        for tipo in ["carro", "moto"]:
            anterior = conn.execute(
                "SELECT valor_hora, valor_dia, minutos_cortesia FROM tarifas WHERE tipo=?", (tipo,)
            ).fetchone()

            vh = float(request.form.get(f"vh_{tipo}", 0))
            vd = float(request.form.get(f"vd_{tipo}", 0))
            mc = int(request.form.get(f"mc_{tipo}", 5))
            conn.execute("UPDATE tarifas SET valor_hora=?, valor_dia=?, minutos_cortesia=? WHERE tipo=?",
                         (vh, vd, mc, tipo))

            # Solo se audita si de verdad cambió algo (evita ruido si el
            # admin solo abrió el formulario y le dio guardar sin tocar nada)
            if anterior and (anterior["valor_hora"] != vh or anterior["valor_dia"] != vd
                              or anterior["minutos_cortesia"] != mc):
                registrar_auditoria(
                    conn,
                    accion="cambiar_tarifa",
                    tabla_afectada="tarifas",
                    registro_id=tipo,
                    valor_anterior=f"hora=${anterior['valor_hora']:,.0f}, dia=${anterior['valor_dia']:,.0f}, cortesia={anterior['minutos_cortesia']}min",
                    valor_nuevo=f"hora=${vh:,.0f}, dia=${vd:,.0f}, cortesia={mc}min",
                    motivo=None,
                )
        conn.commit()
    regs = conn.execute("SELECT * FROM tarifas ORDER BY tipo").fetchall()
    conn.close()
    return render_template("tarifas.html", tarifas=regs)

# ─── Backup ──────────────────────────────────────────────────────
@app.route("/backup")
@login_required
@admin_required
def backup():
    # CAMBIO: Usamos get_colombia_time() para que el nombre del archivo
    # refleje la hora real en la que hiciste la copia.
    ahora_co = get_colombia_time()
    nombre = f"backup_{ahora_co.strftime('%Y%m%d_%H%M%S')}.db"
    
    destino = os.path.join(BACKUP_DIR, nombre)
    shutil.copy2(DB_FILE, destino)
    return send_file(destino, as_attachment=True, download_name=nombre)

# ─── API tiempo real ──────────────────────────────────────────────
@app.route("/api/valor_estimado")
@login_required
def api_valor_estimado():
    placa = request.args.get("placa","").upper()
    conn = get_db()
    reg = conn.execute("SELECT tipo, hora_entrada FROM parqueadero WHERE placa=? AND hora_salida IS NULL ORDER BY id DESC LIMIT 1", (placa,)).fetchone()
    conn.close()
    
    if not reg:
        return jsonify({"error": "No encontrado"})
    
    # 1. Convertimos la entrada y le ponemos la zona horaria de Colombia
    entrada_dt = datetime.strptime(reg["hora_entrada"], "%Y-%m-%d %H:%M:%S")
    tz = pytz.timezone('America/Bogota')
    entrada_dt = tz.localize(entrada_dt)
    
    # 2. Obtenemos la hora actual de Colombia
    ahora_co = get_colombia_time()
    
   # 3. Calculamos los minutos reales (Colombia vs Colombia)
    mins = int(max(0, (ahora_co - entrada_dt).total_seconds() // 60))
    
    h, m = divmod(mins, 60)
    tarifas = get_tarifas()
    valor = int(calcular_valor(reg["tipo"], mins, tarifas))
    
    return jsonify({"tiempo": f"{h}h {m}m", "valor": valor, "mins": mins})

# ─── API para el agente de impresión (DigitalPOS por IP) ─────────
def _check_printer_key():
    if not PRINTER_API_KEY:
        return False
    return hmac.compare_digest(request.headers.get("X-API-KEY", ""), PRINTER_API_KEY)

@app.route("/api/pendientes_impresion")
def api_pendientes_impresion():
    if not _check_printer_key():
        return jsonify({"error": "no autorizado"}), 401

    conn = get_db()
    entradas = conn.execute("""
        SELECT id, placa, tipo, hora_entrada, ticket_num, marca, celular
        FROM parqueadero
        WHERE (impreso_entrada IS NULL OR impreso_entrada = 0)
        ORDER BY id ASC
    """).fetchall()

    salidas = conn.execute("""
        SELECT id, placa, tipo, hora_entrada, hora_salida, valor, ticket_num, metodo_pago
        FROM parqueadero
        WHERE hora_salida IS NOT NULL
        AND (impreso_salida IS NULL OR impreso_salida = 0)
        ORDER BY id ASC
    """).fetchall()
    conn.close()

    return jsonify({
        "entradas": [dict(r) for r in entradas],
        "salidas": [dict(r) for r in salidas],
        "poliza": POLIZA_NUM,
        "certificado": CERTIFICADO_NUM,
        "vigencia": VIGENCIA_POLIZA
    })

@app.route("/api/marcar_impreso/<int:id>/<tipo>", methods=["POST"])
@csrf.exempt  # lo llama el agente local de impresión, autenticado por X-API-KEY, no por sesión de navegador
def api_marcar_impreso(id, tipo):
    if not _check_printer_key():
        return jsonify({"error": "no autorizado"}), 401
    if tipo not in ("entrada", "salida"):
        return jsonify({"error": "tipo invalido"}), 400

    columna = "impreso_entrada" if tipo == "entrada" else "impreso_salida"
    conn = get_db()
    conn.execute(f"UPDATE parqueadero SET {columna} = 1 WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

# ─── Perfil y Seguridad ──────────────────────────────────────────
@app.route("/perfil", methods=["GET", "POST"])
@login_required
def perfil():
    mensaje = None
    if request.method == "POST":
        nueva_pass = request.form.get("password", "").strip()
        confirmar_pass = request.form.get("confirmar_password", "").strip()

        if not nueva_pass:
            mensaje = ("error", "La contraseña no puede estar vacía")
        elif len(nueva_pass) < 6:
            mensaje = ("error", "La contraseña debe tener al menos 6 caracteres")
        elif nueva_pass == confirmar_pass:
            conn = get_db()
            conn.execute("UPDATE usuarios SET password=? WHERE username=?",
                         (generate_password_hash(nueva_pass), session["usuario"]))
            conn.commit()
            conn.close()
            mensaje = ("success", "¡Contraseña actualizada con éxito!")
        else:
            mensaje = ("error", "Las contraseñas no coinciden")
            
    return render_template("perfil.html", mensaje=mensaje)

# ─── Auditoría (Solo Admin) ───────────────────────────────────────
@app.route("/auditoria")
@login_required
@admin_required
def auditoria():
    conn = get_db()
    accion = request.args.get("accion", "").strip()
    if accion:
        regs = conn.execute("""
            SELECT * FROM auditoria WHERE accion = ? ORDER BY id DESC LIMIT 300
        """, (accion,)).fetchall()
    else:
        regs = conn.execute("SELECT * FROM auditoria ORDER BY id DESC LIMIT 300").fetchall()
    acciones = conn.execute("SELECT DISTINCT accion FROM auditoria ORDER BY accion").fetchall()
    conn.close()
    return render_template("auditoria.html", regs=regs, acciones=acciones, accion_filtro=accion)

# ─── Gestión de Usuarios (Solo Admin) ────────────────────────────
@app.route("/usuarios")
@login_required
@admin_required
def gestion_usuarios():
    conn = get_db()
    usuarios = conn.execute("SELECT id, username, rol FROM usuarios").fetchall()
    conn.close()
    return render_template("usuarios.html", usuarios=usuarios)

@app.route("/usuarios/reset/<int:id>", methods=["POST"])
@login_required
@admin_required
def reset_password(id):
    conn = get_db()
    objetivo = conn.execute("SELECT username FROM usuarios WHERE id=?", (id,)).fetchone()

    # Reseteo a clave genérica para que el operador la cambie luego en su perfil
    conn.execute("UPDATE usuarios SET password=? WHERE id=?",
                 (generate_password_hash("cambiar123"), id))

    if objetivo:
        # No se audita ninguna contraseña (ni la vieja ni la nueva) --
        # solo que la acción ocurrió y sobre qué usuario.
        registrar_auditoria(
            conn,
            accion="reset_password",
            tabla_afectada="usuarios",
            registro_id=id,
            valor_nuevo="clave reseteada a valor genérico",
            motivo=f"usuario objetivo: {objetivo['username']}",
        )

    conn.commit()
    conn.close()
    return redirect(url_for('gestion_usuarios'))

# ─── Estadísticas (Solo Admin) ───────────────────────────────────
@app.route("/estadisticas")
@login_required
@admin_required
def estadisticas():
    # 1. IMPORTANTE: Necesitamos importar calendar para el cálculo
    import calendar
    
    conn = get_db()
    
    # 2. CÁLCULO DEL SEMÁFORO (Días restantes del mes)
    ahora = get_colombia_time()
    ultimo_dia = calendar.monthrange(ahora.year, ahora.month)[1]
    dias_restantes = ultimo_dia - ahora.day
    
    # Ingresos últimos 7 días (se excluyen pagos anulados -- no son ingreso real)
    query_semana = """
        SELECT date(hora_salida) as fecha, SUM(valor) as total
        FROM parqueadero
        WHERE hora_salida IS NOT NULL
        AND date(hora_salida) > date('now', '-7 days')
        AND (anulado = 0 OR anulado IS NULL)
        GROUP BY date(hora_salida)
        ORDER BY fecha ASC
    """
    datos = conn.execute(query_semana).fetchall()

    # Ingresos por tipo (idem, sin pagos anulados)
    por_tipo = conn.execute("""
        SELECT tipo, SUM(valor) as total
        FROM parqueadero
        WHERE hora_salida IS NOT NULL
        AND (anulado = 0 OR anulado IS NULL)
        GROUP BY tipo
    """).fetchall()
    conn.close()

    labels = [d['fecha'] for d in datos]
    valores = [d['total'] for d in datos]

    # 3. PASAR dias_restantes AL TEMPLATE
    return render_template("estadisticas.html",
                           labels=labels,
                           valores=valores,
                           por_tipo=por_tipo,
                           dias_restantes=dias_restantes)

# ─── Limpiar Pruebas (Solo Admin, y solo en entorno de desarrollo) ─
# Borra TODO el histórico de parqueadero. Se bloquea por completo a
# menos que la variable de entorno DEVELOPMENT=true esté activa, para
# que nunca se pueda disparar por accidente en producción (Railway).
@app.route("/admin/limpiar_pruebas", methods=["POST"])
@login_required
@admin_required
def limpiar_pruebas():
    if os.environ.get("DEVELOPMENT") != "true":
        return "Esta función solo está disponible en entorno de desarrollo.", 403
    conn = get_db()
    conn.execute("DELETE FROM parqueadero")
    conn.execute("UPDATE consecutivo SET numero=0 WHERE id=1")
    conn.commit()
    conn.close()
    return redirect(url_for('inicio'))

@app.route("/imprimir_caja")
@login_required
def imprimir_caja():
    ahora_co = get_colombia_time()
    fecha_hoy = ahora_co.strftime("%Y-%m-%d")
    
    conn = get_db()
    # 1. Obtenemos todos los registros que salieron hoy (sin anulados:
    #    un pago anulado no es un ingreso real, no debe sumar en el cierre)
    regs = conn.execute("""
        SELECT metodo_pago, valor, tipo
        FROM parqueadero
        WHERE date(hora_salida) = ?
          AND (anulado = 0 OR anulado IS NULL)
    """, (fecha_hoy,)).fetchall()
    
    # 2. Calculamos totales y desglose
    total_caja = 0
    total_entradas_hoy = conn.execute("SELECT COUNT(*) FROM parqueadero WHERE date(hora_entrada) = ?", (fecha_hoy,)).fetchone()[0]
    total_salidas_hoy = len(regs)
    
    desglose = {}
    for r in regs:
        valor = r["valor"] or 0
        total_caja += valor
        # Limpiamos el nombre del método (quitamos convenios si existen)
        m = (r["metodo_pago"] or "Efectivo").split(" - ")[0]
        desglose[m] = desglose.get(m, 0) + valor
        
    conn.close()
    
    return render_template("ticket_caja.html", 
                           fecha=ahora_co.strftime("%d/%m/%Y"),
                           hora=ahora_co.strftime("%I:%M %p"),
                           desglose=desglose,
                           total_caja=int(total_caja),
                           total_entradas=total_entradas_hoy,
                           total_salidas=total_salidas_hoy)

# ─── Ruta para Anular Pago ─────────────────────────────────────
@app.route("/anular_pago/<int:id>", methods=["POST"])
@login_required
@admin_required
def anular_pago(id):

    # Capturamos los datos del formulario (antes venían por la URL/GET,
    # lo que permitía anular un pago con solo hacer clic en un link).
    motivo = request.form.get("motivo", "No especificado").strip()
    valor_real_raw = request.form.get("real", "0")

    # Convertimos el valor a pesos enteros
    try:
        valor_real = int(
            str(valor_real_raw)
            .replace("$", "")
            .replace(".", "")
            .replace(",", "")
            .strip()
        )
    except (ValueError, TypeError):
        valor_real = 0

    conn = get_db()

    # Verificamos que exista y obtenemos estado de anulación + el valor
    # actual (lo necesitamos para el registro de auditoría, y sobre todo
    # para NO perderlo: antes esta ruta hacía "valor = 0" y el cobro
    # original quedaba destruido para siempre, sin forma de saber cuánto
    # se había cobrado antes de la anulación).
    reg = conn.execute("""
        SELECT id, anulado, valor
        FROM parqueadero
        WHERE id = ?
    """, (id,)).fetchone()

    if reg:

        # Evitar doble anulación
        if reg["anulado"]:
            conn.close()
            flash("Este pago ya estaba anulado.", "error")
            return redirect(url_for("historico"))

        valor_original = reg["valor"]

        # Registrar anulación.
        # IMPORTANTE: ya NO se pone "valor = 0". El cobro original se
        # conserva intacto en el histórico; el flag `anulado` es lo que
        # lo excluye de caja/estadísticas (ver esas rutas), y
        # `valor_real` guarda lo que el admin indica que realmente se
        # cobró (si acaso se cobró algo).
        conn.execute("""
            UPDATE parqueadero
            SET
                anulado = 1,
                motivo_anulacion = ?,
                valor_real = ?
            WHERE id = ?
        """, (
            motivo,
            valor_real,
            id
        ))

        registrar_auditoria(
            conn,
            accion="anular_pago",
            tabla_afectada="parqueadero",
            registro_id=id,
            valor_anterior=f"valor original: ${valor_original:,.0f}",
            valor_nuevo=f"anulado; valor real indicado: ${valor_real:,.0f}",
            motivo=motivo,
        )

        conn.commit()

    conn.close()

    return redirect(url_for("historico"))
    
# Mueve la llamada afuera del if para que Railway la ejecute sí o sí
init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
